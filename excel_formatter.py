"""
Formatters for Excel
"""
from typing import List, Optional, Tuple
import copy

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from helpers import ShowData

X_PADDING = 0.7
Y_PADDING = 0.7
Y_PADDING_HEADER = 0.4
HEAD_FOOT_PAD = 0.2

PAGE_HEIGHT_INCHES = 11


def page_setup(ws: Worksheet, rows_to_repeat: int = 0) -> None:
    """
    Set the page size, margins, and default view.
    """
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_options.horizontalCentered = True
    ws.sheet_view.view = "pageLayout"

    ws.page_margins = PageMargins(
        bottom=Y_PADDING,
        top=Y_PADDING + Y_PADDING_HEADER,
        left=X_PADDING,
        right=X_PADDING,
        header=HEAD_FOOT_PAD,
        footer=HEAD_FOOT_PAD,
    )
    ws.sheet_view.showGridLines = False

    # Force the title + column names to repeat
    if rows_to_repeat > 0:
        ws.print_title_rows = f"1:{rows_to_repeat}"


def add_title(ws: Worksheet, name: str, show_info: Optional[ShowData] = None) -> None:
    """
    Add header and footer to worksheet
    """
    if ws.oddHeader is None:
        raise RuntimeError("oddHeader is not writable!")
    if show_info is not None:
        # Header
        ws.oddHeader.left.text = f"&[Date]\n {show_info.revision}"
        ws.oddHeader.left.size = 12
        ws.oddHeader.right.text = f"{show_info.show_name}\nLD: {show_info.ld_name}"
        ws.oddHeader.right.size = 12
    else:
        ws.oddHeader.left.text = "&[Date]"
        ws.oddHeader.left.size = 12

    # Title
    ws.oddHeader.center.text = name
    ws.oddHeader.center.size = 22
    ws.oddHeader.center.font = "Calibri,Bold"

    # Footer
    if ws.oddFooter is None:
        raise RuntimeError("oddFooter is not writable!")
    ws.oddFooter.left.text = name
    ws.oddFooter.left.size = 12
    ws.oddFooter.right.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.right.size = 12


def set_col_widths(ws: Worksheet, width: List[int], page_width: int) -> None:
    """
    Set the widths of a page in terms of % of a full page

    Widths are provided in terms of percentages, but excel expects px
    Assume page width is 610px (experimentally derived)
    I think it's 96 ppi so 96 * usable page width?
    https://www.reddit.com/r/excel/comments/l9k99z/why_does_excel_use_different_units_of_measurement/
    """
    width_px = [w * 0.01 * 610 * (page_width / 100) for w in width]

    for i in range(1, ws.max_column + 1):
        # why tf do we divide by 7
        ws.column_dimensions[get_column_letter(i)].width = int(width_px[i - 1] / 7)


def wrap_all_cells(ws: Worksheet) -> None:
    """
    Force all cells to wrap text instead of overflow. 
    """
    for row in ws.iter_rows():
        for cell in row:
            alignment = copy.copy(cell.alignment)
            alignment.wrapText = True
            cell.alignment = alignment
