from typing import List, Optional, Tuple

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from helpers import ShowData


def instrschedule_cols(ws: Worksheet, start_row: int) -> None:
    # Format the header
    ws.merge_cells(
        start_row=start_row, start_column=1, end_row=start_row, end_column=ws.max_column
    )
    header_cell = ws.cell(start_row, 1)
    header_cell.font = Font(size=18, bold=True)
    header_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Format the column labels
    side = Side(border_style="thin", color="FF000000")
    for i, cell in enumerate(ws[f"{start_row + 1}:{start_row + 1}"]):
        if i == 0:
            # Center the U# column
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

        cell.font = Font(size=12, bold=True)
        cell.border = Border(bottom=side, top=side)

    # Format data
    for row in ws.iter_rows(min_row=start_row + 2, max_row=ws.max_row):
        # Add dashed line if next row isn't the same U#
        side = Side(border_style="dashed", color="FF000000")
        if row[0].value != ws.cell(row[0].row + 1, 1).value:
            add_line = True
        else:
            add_line = False

        for idx, cell in enumerate(row):
            if idx == 0:
                # Center the U# column
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif idx == 4:
                # Channel formatting
                cell.number_format = "(0)"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif idx == 5:
                # Address formatting
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

            cell.font = Font(size=11)
            if add_line:
                cell.border = Border(bottom=side)


def colorcut_cols(ws: Worksheet) -> None:
    side = Side(border_style="thin", color="FF000000")
    for i, cell in enumerate(ws["1:1"]):
        if i == 0:
            # Center color
            cell.alignment = Alignment(horizontal="center", vertical="bottom")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="bottom")
        cell.font = Font(size=14, bold=True)
        cell.border = Border(bottom=side)

    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            if idx == 0:
                # Gel formatting
                cell.font = Font(size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(size=14)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

    # TODO: add bar on last row


def gobos(ws: Worksheet) -> None:
    side = Side(border_style="thin", color="FF000000")
    for i, cell in enumerate(ws["1:1"]):
        if i == 0:
            # Center color
            cell.alignment = Alignment(horizontal="center", vertical="bottom")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="bottom")
        cell.font = Font(size=14, bold=True)
        cell.border = Border(bottom=side)

    ws.row_dimensions[1].height = 30

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row):
            if idx == 0:
                # Gobo formatting
                cell.font = Font(size=14)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = Font(size=14)
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

    # TODO: add bar on last row


def repeated_gels(ws: Worksheet) -> None:
    side = Side(border_style="thin", color="FF000000")

    for row in range(2, ws.max_row):
        # Search forward until non-matching row found
        end_row = row
        max_row = ws.max_row
        while (
            end_row <= max_row
            and ws.cell(end_row + 1, 1).value == ws.cell(row, 1).value
        ):
            end_row += 1

        if row != end_row:
            # Merge cells
            ws.merge_cells(start_row=row, end_row=end_row, start_column=1, end_column=1)

        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).border = Border(top=side)

        row = end_row - 1


def repeated_channel(ws: Worksheet) -> None:
    prev_channel = 0
    prev_row = None
    side = Side(border_style="thin", color="FF000000")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].value == prev_channel:
            # This is a repeated channel!
            row[0].value = ""
            for idx, cell in enumerate(row):
                # if value the same as the last, don't repeat the data
                # exception: do repeat U# since this could get confusing
                if cell.value == prev_row[idx].value and idx != 3:
                    if prev_row[idx].value != "":
                        # Don't "-ify empty fields
                        cell.value = '"'
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        else:
            prev_channel = row[0].value
            prev_row = row
            for idx, cell in enumerate(row):
                cell.border = Border(top=side)


# Good heavens this is hacky
def instr_schedule_pagebreaks(ws: Worksheet) -> None:
    # Goal: each position should fit on a page (or at least take up a full page otherwise)
    pos_start_index = 0
    last_height = 0
    cur_height = 0.0

    # Note: all math here done in inches. it's hacky but also excel sucks so
    PAGE_FUDGE = 0.7  # Adjust this if you have weird overflow issues.
    PAGE_HEIGHT = (PAGE_HEIGHT_INCHES - (Y_PADDING * 2 + Y_PADDING_HEADER)) - PAGE_FUDGE

    TYPE_LINEBREAK_LEN = 30
    COLOR_LINEBREAK_LEN = 25

    for row in range(2, ws.max_row):
        # calculate how long this position is
        if ws.cell(row, 5).value is None and ws.cell(row, 1).value is None:
            # end of position
            if last_height + cur_height > PAGE_HEIGHT:
                # we don't want to add this position to the same page, add pagebreak
                ws.row_breaks.append(Break(id=pos_start_index - 1))
                last_height = cur_height
            else:
                last_height = last_height + cur_height + 0.22

            cur_height = 0
        else:
            # Predict height based on value
            if (not ws.cell(row, 1).value.isdigit()) and ws.cell(row, 2).value is None:
                # Position names have heights of 0.33"
                cur_height += 0.33
                pos_start_index = row
            elif ws.cell(row, 1).value.isdigit() or ws.cell(row, 1).value == "U#":
                # Probably a U#, which has default height 0.22"
                if (
                    len(ws.cell(row, 3).value) > TYPE_LINEBREAK_LEN
                    or len(ws.cell(row, 4).value) > COLOR_LINEBREAK_LEN
                ):
                    # Double height
                    cur_height += 0.44
                else:
                    cur_height += 0.22
            else:
                # dunno, assume it's a standard row
                cur_height += 0.22
