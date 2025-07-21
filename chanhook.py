"""
Generate lighting export by going to File > Export > Export Lighting Device Data...
Select all fields and then export. Output should be a txt file.
"""

# TODO: fix same color merge column, remove empty channel
# TODO: format numbers as numbers in excel
# TODO: accessories???
# TODO: embed UUID into a hidden cell somehow
# TODO: Add a field in VW as a "rep" indicator, then use that on color hookup to determine existing vs new
# TODO: add title page with links
# TODO: make setting widths relative to 100% page
import argparse
import logging
import os
import re
from dataclasses import dataclass
from typing import List, Optional, Tuple

import pandas as pd
from natsort import natsort_keygen, natsorted
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from pandas.io.formats.style import Styler

import excel_format
import generate_df
from channel_hookup import ChannelHookup
from color_cut_list import ColorCutList
from helpers import FontStyle, ShowData
from instrument_schedule import InstrumentSchedule
from vectorworks_xml import VWExport


def is_file(path: str) -> str:
    if not os.path.isfile(path):
        raise argparse.ArgumentTypeError("Path is not a valid file")

    return path


def add_colorcuts(wb: Workbook, vw_export: pd.DataFrame, show_info: ShowData) -> None:
    colors = generate_df.colorcut(vw_export)

    # Dump to worksheet
    ws = wb.create_sheet("Color Cut List", -1)
    for r in dataframe_to_rows(colors, index=False, header=True):
        ws.append(r)

    # Format
    excel_format.add_title(ws, "Color Cut List", show_info)
    excel_format.page_setup(ws, 1)
    excel_format.set_col_widths(ws, [80, 100, 50])
    excel_format.repeated_gels(ws)
    excel_format.colorcut_cols(ws)


def add_gobos(wb: Workbook, vw_export: pd.DataFrame, show_info: ShowData) -> None:
    gobo_list = generate_df.gobo_pull(vw_export)

    # Dump to Worksheet
    ws = wb.create_sheet("Gobo Pull List", -1)
    ws.append(["Gobo", "Qty"])
    for r in dataframe_to_rows(gobo_list, index=False, header=False):
        ws.append(r)

    # Format
    excel_format.add_title(ws, "Gobo Pull List", show_info)
    excel_format.page_setup(ws, 1)
    excel_format.set_col_widths(ws, [200, 50])
    excel_format.gobos(ws)


def add_channel_hookup_html(vw_export, show_info):
    chans = ColorCutList(vw_export, show_info)
    html = chans.make()
    with open("chans.html", "w") as f:
        f.write(html)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("raw", help="Raw instrument from Vectorworks", type=is_file)
    parser.add_argument("--show", help="Show name")
    parser.add_argument("--ld", help="Lighting designer initials")
    parser.add_argument("--rev", help="Revision string (ex. 'Rev. A')")

    args = parser.parse_args()
    logging.basicConfig(level=logging.DEBUG)

    show_info = ShowData(args.show, args.ld, args.rev)

    if "csv" in args.raw:
        # Converter is to supress the warning when I set addr=0 to empty string
        vw_export = pd.read_csv(
            args.raw, sep="\t", header=0, converters={"Absolute Address": str}
        )

        # Clear VW's default "None" character
        vw_export = vw_export.replace("-", "")

    elif "xml" in args.raw:
        vw_export = VWExport(args.raw).export_df()

    wb = Workbook()

    add_channel_hookup_html(vw_export, show_info)
    return
    add_channel_hookup(wb, vw_export, show_info)
    add_instrument_schedule(wb, vw_export, show_info)
    add_colorcuts(wb, vw_export, show_info)
    add_gobos(wb, vw_export, show_info)

    del wb["Sheet"]
    output_filename = (
        f"{show_info.show_name.replace(' ', '')}_Paperwork_"
        + re.sub(r"\W+", "", show_info.revision)
        + ".xlsx"
    )
    wb.save(output_filename)


if __name__ == "__main__":
    main()
